import streamlit as st
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Excel 上傳→修改→下載", page_icon="📄", layout="centered")
st.title("📄 Excel 雲端修改器（上傳→修改→下載）")

uploaded = st.file_uploader("請上傳 .xlsx", type=["xlsx"])
st.caption("＊檔案僅在記憶體處理，不會長期存放伺服器。")

#（示例）提供一組可寫入的欄位；你可依需求調整
with st.expander("選填：要追加的一列資料"):
    date_str = st.text_input("日期（以『字串』寫入，例如 20250819）", "")
    c1,c2,c3 = st.columns(3)
    with c1:
        v1 = st.number_input("Value_1", value=0.0, step=1.0, format="%.3f")
        v2 = st.number_input("Value_2", value=0.0, step=1.0, format="%.3f")
    with c2:
        v3 = st.number_input("Value_3", value=0.0, step=1.0, format="%.3f")
        v4 = st.number_input("Value_4", value=0.0, step=1.0, format="%.3f")
    with c3:
        v5 = st.number_input("Value_5", value=0.0, step=1.0, format="%.3f")
        v6 = st.number_input("Value_6", value=0.0, step=1.0, format="%.3f")
    note = st.text_input("備註", "")

target_sheet = st.text_input("要寫入的工作表名稱（預設：Data）", value="Data")
add_timestamp = st.checkbox("下載檔名加上時間戳", value=False)

def append_row(ws, values):
    """將資料追加到該工作表的下一列（自動建立表頭）。"""
    if ws.max_row == 1 and all((cell.value is None) for cell in ws[1]):
        # 空表，建表頭
        ws.append(["date_str","value_1","value_2","value_3","value_4","value_5","value_6","note"])
    ws.append(values)
    # 讓第一欄為字串格式（避免被 Excel 自動轉日期）
    col1 = get_column_letter(1)
    for cell in ws[f"{col1}1":f"{col1}{ws.max_row}"][0]:
        cell.number_format = "@"

if uploaded is not None:
    st.success(f"已上傳：{uploaded.name}")
    if st.button("開始修改並提供下載"):
        # 讀入到 openpyxl
        data = uploaded.read()
        wb = load_workbook(BytesIO(data), data_only=False, keep_vba=False)

        # 取得/建立目標工作表
        ws = wb[target_sheet] if target_sheet in wb.sheetnames else wb.create_sheet(title=target_sheet)

        # 若有輸入日期（代表要追加一列）
        if date_str:
            if not (len(date_str) == 8 and date_str.isdigit()):
                st.error("日期需為 8 位數字（YYYYMMDD）。")
                st.stop()
            append_row(ws, [date_str, v1, v2, v3, v4, v5, v6, note])

        # 這裡也可以做「指定儲存格改值」的客製化：
        # ws["B2"].value = "Hello"  # 範例

        # 轉回位元串供下載
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        # 檔名
        base = uploaded.name.rsplit(".xlsx", 1)[0]
        if add_timestamp:
            ts = datetime.now().strftime("%Y%m%d-%H%M%S")
            fname = f"{base}-{ts}.xlsx"
        else:
            fname = f"{base}.xlsx"

        st.download_button(
            label="📥 下載修改後的 Excel",
            data=out.getvalue(),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.info("若你的圖表來源綁『表格 (Ctrl+T)』或動態範圍，追加資料後打開檔案圖表會自動延伸。")
else:
    st.info("請先上傳 .xlsx 檔。")